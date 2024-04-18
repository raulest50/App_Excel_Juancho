

from openpyxl import load_workbook, Workbook

# Loading the workbook
wb = load_workbook('../mkup_excel.xlsx')


def prepare_workbook(workbook):
    # si hay mas de 1 hoja se eliminan, en caso contrario no se hace nada
    sheets = workbook.sheetnames
    if not len(sheets) == 1:
        sheets2delete = sheets[1:] ## todas las hojas excepto la primera (index 0)
        print(f"sheets to delete: {sheets2delete}")
        for s2d in sheets2delete:
            del workbook[s2d]
    mdh_s = wb.create_sheet(title="MDH")
    tin_s = wb.create_sheet(title="TIN")
    like_s = wb.create_sheet(title="LIKE")
    alma_s = wb.create_sheet(title="ALMA")
    sheets = [mdh_s, tin_s, like_s, alma_s]
    return workbook, sheets


def resize_ph_cols(workbook, sheets):
    size = 30
    for sh in sheets:
        sh.column_dimensions['C'].width = size
        sh.column_dimensions['D'].width = size
        sh.column_dimensions['E'].width = size
        sh.column_dimensions['F'].width = size
        sh.column_dimensions['G'].width = size
    return workbook, sheets


wb, sheets = prepare_workbook(workbook=wb)
wb, sheets = resize_ph_cols(workbook=wb, sheets=sheets)
wb.save('mkup_excel.xlsx')

# Selecting a worksheet
main_ws = wb['Despacho']

finished = False
row = 4
bodega = ""
while not finished:
    bodega = main_ws[f"A{row}"]
    match bodega:
        case "MDH":
            pass
        case "TIN":
            pass
        case "LIKE":
            pass
        case "ALMA BEAUTY":
            pass
        case _:
            finished = True
    row = row + 1
print(row)


# Modifying a cell's value
#ws['A1'] = 42

# Saving the modified workbook
#wb.save('modified_file.xlsx')
